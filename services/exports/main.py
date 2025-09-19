from fastapi import FastAPI, Depends, HTTPException, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional, Dict
import pandas as pd
import io
import base64
import json
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO, StringIO
import aiofiles
from pathlib import Path

from shared.database import get_db
from shared.models import User, Department, Class, Exam, Mark, Question, AuditLog
from shared.auth import RoleChecker
from shared.schemas import UserResponse, DepartmentResponse, ClassResponse, ExamResponse, MarkResponse

app = FastAPI(title="Export Service", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

def log_audit(db: Session, user_id: int, action: str, table_name: str, details: str = None):
    """Log audit trail"""
    import json
    
    # Format details as JSON string
    details_json = json.dumps({
        "message": details or f"{action} on {table_name}",
        "timestamp": datetime.utcnow().isoformat()
    })
    
    audit_log = AuditLog(
        user_id=user_id,
        action=action,
        table_name=table_name,
        old_values=None,
        new_values=details_json,
        created_at=datetime.utcnow()
    )
    db.add(audit_log)
    db.commit()

async def export_to_csv(df: pd.DataFrame, data_type: str):
    """Export DataFrame to CSV"""
    csv_buffer = StringIO()
    df.to_csv(csv_buffer, index=False)
    csv_content = csv_buffer.getvalue()
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={data_type}_export.csv"}
    )

async def export_to_pdf(df: pd.DataFrame, title: str, data_type: str):
    """Export DataFrame to PDF"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_para = Paragraph(title, title_style)
        elements.append(title_para)
        elements.append(Spacer(1, 12))
        
        # Convert DataFrame to table data
        table_data = [df.columns.tolist()] + df.values.tolist()
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        pdf_content = buffer.getvalue()
        buffer.close()
        
        return Response(
            content=pdf_content,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={data_type}_export.pdf"}
        )
    except ImportError:
        # Fallback to CSV if reportlab is not available
        return await export_to_csv(df, data_type)

async def export_to_excel(df: pd.DataFrame, title: str, data_type: str):
    """Export DataFrame to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        import io
        
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()
        
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={data_type}_export.xlsx"}
        )
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return await export_to_csv(df, data_type)

def create_chart(data: pd.DataFrame, chart_type: str, title: str) -> str:
    """Create a chart and return as base64 string"""
    plt.figure(figsize=(10, 6))
    
    if chart_type == 'bar':
        if len(data.columns) >= 2:
            plt.bar(data.iloc[:, 0], data.iloc[:, 1])
            plt.xlabel(data.columns[0])
            plt.ylabel(data.columns[1])
    elif chart_type == 'pie':
        if len(data.columns) >= 2:
            plt.pie(data.iloc[:, 1], labels=data.iloc[:, 0], autopct='%1.1f%%')
    elif chart_type == 'line':
        if len(data.columns) >= 2:
            plt.plot(data.iloc[:, 0], data.iloc[:, 1], marker='o')
            plt.xlabel(data.columns[0])
            plt.ylabel(data.columns[1])
    
    plt.title(title)
    plt.tight_layout()
    
    # Save to base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png', bbox_inches='tight', dpi=300)
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    
    return image_base64

# Root endpoint
@app.get("/")
async def root():
    return {"message": "Export Service is running", "status": "healthy"}

# Export endpoints
@app.get("/api/exports/users/csv")
async def export_users_csv_full_path(
    format: str = Query("csv", description="Export format: csv, pdf, excel"),
    role: Optional[str] = Query(None),
    department_id: Optional[int] = Query(None),
    is_active: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(RoleChecker(["admin", "hod", "teacher"]))
):
    """Export users in specified format"""
    return await export_users(role, department_id, is_active, db, current_user_id, format)

async def export_users(
    role: Optional[str],
    department_id: Optional[int],
    is_active: Optional[bool],
    db: Session,
    current_user_id: int,
    format: str = "csv"
):
    """Export users in specified format (csv, pdf, excel)"""
    current_user = db.query(User).filter(User.id == current_user_id).first()
    
    # Build query
    query = db.query(User).options(
        joinedload(User.department),
        joinedload(User.class_assigned)
    )
    
    # Apply role-based filters
    if current_user.role == "hod":
        query = query.filter(User.department_id == current_user.department_id)
    
    # Apply filters
    if role:
        query = query.filter(User.role == role)
    if department_id:
        query = query.filter(User.department_id == department_id)
    if is_active is not None:
        query = query.filter(User.is_active == is_active)
    
    users = query.all()
    
    # Prepare data
    users_data = []
    for user in users:
        users_data.append({
            'ID': user.id,
            'Username': user.username,
            'Email': user.email,
            'Full Name': user.full_name,
            'Role': user.role,
            'Phone': user.phone or '',
            'Address': user.address or '',
            'Department': user.department.name if user.department else '',
            'Class': user.class_assigned.name if user.class_assigned else '',
            'Student ID': user.student_id or '',
            'Employee ID': user.employee_id or '',
            'Active': user.is_active,
            'Created At': user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '',
            'Last Login': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else ''
        })
    
    df = pd.DataFrame(users_data)
    
    # Log audit trail
    log_audit(db, current_user_id, "export", "users", f"Exported {len(users)} users to {format.upper()}")
    
    if format.lower() == "pdf":
        return await export_to_pdf(df, "Users Export", "users")
    elif format.lower() in ["excel", "xlsx"]:
        return await export_to_excel(df, "Users Export", "users")
    else:  # Default to CSV
        return await export_to_csv(df, "users")

@app.get("/users/csv")
async def export_users_csv(
    role: Optional[str] = Query(None),
    department_id: Optional[int] = Query(None),
    is_active: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(RoleChecker(["admin", "hod"]))
):
    """Export users to CSV"""
    current_user = db.query(User).filter(User.id == current_user_id).first()
    
    # Build query
    query = db.query(User).options(
        joinedload(User.department),
        joinedload(User.class_assigned)
    )
    
    # Apply role-based filters
    if current_user.role == "hod":
        query = query.filter(User.department_id == current_user.department_id)
    
    # Apply additional filters
    if role:
        query = query.filter(User.role == role)
    if department_id:
        query = query.filter(User.department_id == department_id)
    if is_active is not None:
        query = query.filter(User.is_active == is_active)
    
    users = query.all()
    
    # Convert to DataFrame
    users_data = []
    for user in users:
        users_data.append({
            'ID': user.id,
            'Username': user.username,
            'Email': user.email,
            'Full Name': user.full_name,
            'Role': user.role,
            'Phone': user.phone or '',
            'Address': user.address or '',
            'Department': user.department.name if user.department else '',
            'Class': user.class_assigned.name if user.class_assigned else '',
            'Student ID': user.student_id or '',
            'Employee ID': user.employee_id or '',
            'Active': user.is_active,
            'Created At': user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '',
            'Last Login': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else ''
        })
    
    df = pd.DataFrame(users_data)
    
    # Create CSV
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False)
    csv_content = csv_buffer.getvalue()
    
    # Log the export
    log_audit(db, current_user_id, "export", "users", f"Exported {len(users)} users to CSV")
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=users_export.csv"}
    )

@app.get("/departments/csv")
async def export_departments_csv(
    db: Session = Depends(get_db),
    current_user_id: int = Depends(RoleChecker(["admin", "hod"]))
):
    """Export departments to CSV"""
    current_user = db.query(User).filter(User.id == current_user_id).first()
    
    query = db.query(Department)
    
    # HOD can only see their own department
    if current_user.role == "hod":
        query = query.filter(Department.id == current_user.department_id)
    
    departments = query.all()
    
    # Convert to DataFrame
    departments_data = []
    for dept in departments:
        departments_data.append({
            'ID': dept.id,
            'Name': dept.name,
            'Code': dept.code,
            'Description': dept.description or '',
            'Head of Department': dept.hod_name or '',
            'Email': dept.email or '',
            'Phone': dept.phone or '',
            'Created At': dept.created_at.strftime('%Y-%m-%d %H:%M:%S') if dept.created_at else ''
        })
    
    df = pd.DataFrame(departments_data)
    
    # Create CSV
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False)
    csv_content = csv_buffer.getvalue()
    
    # Log the export
    log_audit(db, current_user_id, "export", "departments", f"Exported {len(departments)} departments to CSV")
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=departments_export.csv"}
    )

@app.get("/classes/csv")
async def export_classes_csv(
    department_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(RoleChecker(["admin", "hod", "teacher"]))
):
    """Export classes to CSV"""
    current_user = db.query(User).filter(User.id == current_user_id).first()
    
    query = db.query(Class).options(joinedload(Class.department))
    
    # Apply role-based filters
    if current_user.role == "hod":
        query = query.filter(Class.department_id == current_user.department_id)
    elif current_user.role == "teacher":
        query = query.filter(Class.teacher_id == current_user_id)
    
    # Apply additional filters
    if department_id:
        query = query.filter(Class.department_id == department_id)
    
    classes = query.all()
    
    # Convert to DataFrame
    classes_data = []
    for cls in classes:
        classes_data.append({
            'ID': cls.id,
            'Name': cls.name,
            'Code': cls.code,
            'Department': cls.department.name if cls.department else '',
            'Teacher': cls.teacher.full_name if cls.teacher else '',
            'Subject': cls.subject or '',
            'Academic Year': cls.academic_year or '',
            'Semester': cls.semester or '',
            'Created At': cls.created_at.strftime('%Y-%m-%d %H:%M:%S') if cls.created_at else ''
        })
    
    df = pd.DataFrame(classes_data)
    
    # Create CSV
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False)
    csv_content = csv_buffer.getvalue()
    
    # Log the export
    log_audit(db, current_user_id, "export", "classes", f"Exported {len(classes)} classes to CSV")
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=classes_export.csv"}
    )

@app.get("/exams/csv")
async def export_exams_csv(
    class_id: Optional[int] = Query(None),
    exam_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(RoleChecker(["admin", "hod", "teacher"]))
):
    """Export exams to CSV"""
    current_user = db.query(User).filter(User.id == current_user_id).first()
    
    query = db.query(Exam).options(joinedload(Exam.class_assigned))
    
    # Apply role-based filters
    if current_user.role == "hod":
        query = query.filter(Exam.class_assigned.has(Class.department_id == current_user.department_id))
    elif current_user.role == "teacher":
        query = query.filter(Exam.class_assigned.has(Class.teacher_id == current_user_id))
    
    # Apply additional filters
    if class_id:
        query = query.filter(Exam.class_id == class_id)
    if exam_type:
        query = query.filter(Exam.exam_type == exam_type)
    
    exams = query.all()
    
    # Convert to DataFrame
    exams_data = []
    for exam in exams:
        exams_data.append({
            'ID': exam.id,
            'Title': exam.title,
            'Description': exam.description or '',
            'Class': exam.class_assigned.name if exam.class_assigned else '',
            'Type': exam.exam_type,
            'Duration': exam.duration_minutes,
            'Total Marks': exam.total_marks,
            'Passing Marks': exam.passing_marks,
            'Start Time': exam.start_time.strftime('%Y-%m-%d %H:%M:%S') if exam.start_time else '',
            'End Time': exam.end_time.strftime('%Y-%m-%d %H:%M:%S') if exam.end_time else '',
            'Created At': exam.created_at.strftime('%Y-%m-%d %H:%M:%S') if exam.created_at else ''
        })
    
    df = pd.DataFrame(exams_data)
    
    # Create CSV
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False)
    csv_content = csv_buffer.getvalue()
    
    # Log the export
    log_audit(db, current_user_id, "export", "exams", f"Exported {len(exams)} exams to CSV")
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=exams_export.csv"}
    )

@app.get("/marks/csv")
async def export_marks_csv(
    exam_id: Optional[int] = Query(None),
    class_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(RoleChecker(["admin", "hod", "teacher"]))
):
    """Export marks to CSV"""
    current_user = db.query(User).filter(User.id == current_user_id).first()
    
    query = db.query(Mark).options(
        joinedload(Mark.student),
        joinedload(Mark.exam)
    )
    
    # Apply role-based filters
    if current_user.role == "hod":
        query = query.filter(Mark.exam.has(Exam.class_assigned.has(Class.department_id == current_user.department_id)))
    elif current_user.role == "teacher":
        query = query.filter(Mark.exam.has(Exam.class_assigned.has(Class.teacher_id == current_user_id)))
    
    # Apply additional filters
    if exam_id:
        query = query.filter(Mark.exam_id == exam_id)
    if class_id:
        query = query.filter(Mark.exam.has(Exam.class_id == class_id))
    
    marks = query.all()
    
    # Convert to DataFrame
    marks_data = []
    for mark in marks:
        marks_data.append({
            'ID': mark.id,
            'Student': mark.student.full_name if mark.student else '',
            'Student ID': mark.student.student_id if mark.student else '',
            'Exam': mark.exam.title if mark.exam else '',
            'Class': mark.exam.class_assigned.name if mark.exam and mark.exam.class_assigned else '',
            'Marks Obtained': mark.marks_obtained,
            'Total Marks': mark.exam.total_marks if mark.exam else 0,
            'Percentage': (mark.marks_obtained / mark.exam.total_marks * 100) if mark.exam and mark.exam.total_marks > 0 else 0,
            'Grade': mark.grade or '',
            'Status': mark.status,
            'Submitted At': mark.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if mark.submitted_at else '',
            'Created At': mark.created_at.strftime('%Y-%m-%d %H:%M:%S') if mark.created_at else ''
        })
    
    df = pd.DataFrame(marks_data)
    
    # Create CSV
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False)
    csv_content = csv_buffer.getvalue()
    
    # Log the export
    log_audit(db, current_user_id, "export", "marks", f"Exported {len(marks)} marks to CSV")
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=marks_export.csv"}
    )

@app.get("/charts/bar")
async def create_bar_chart(
    data_type: str = Query(..., description="Type of data to chart"),
    class_id: Optional[int] = Query(None),
    exam_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(RoleChecker(["admin", "hod", "teacher"]))
):
    """Create a bar chart"""
    current_user = db.query(User).filter(User.id == current_user_id).first()
    
    # Prepare data based on type
    if data_type == "marks_distribution":
        query = db.query(Mark).options(joinedload(Mark.exam))
        if exam_id:
            query = query.filter(Mark.exam_id == exam_id)
        elif class_id:
            query = query.filter(Mark.exam.has(Exam.class_id == class_id))
        
        marks = query.all()
        data = pd.DataFrame([{
            'Grade': mark.grade or 'Ungraded',
            'Count': 1
        } for mark in marks])
        
        if not data.empty:
            data = data.groupby('Grade').size().reset_index(name='Count')
            chart_base64 = create_chart(data, 'bar', 'Marks Distribution')
        else:
            chart_base64 = None
    
    elif data_type == "exam_performance":
        query = db.query(Exam).options(joinedload(Exam.class_assigned))
        if class_id:
            query = query.filter(Exam.class_id == class_id)
        
        exams = query.all()
        data = pd.DataFrame([{
            'Exam': exam.title,
            'Average Marks': db.query(Mark).filter(Mark.exam_id == exam.id).with_entities(
                db.func.avg(Mark.marks_obtained)
            ).scalar() or 0
        } for exam in exams])
        
        if not data.empty:
            chart_base64 = create_chart(data, 'bar', 'Exam Performance')
        else:
            chart_base64 = None
    
    else:
        raise HTTPException(status_code=400, detail="Invalid data type")
    
    if chart_base64:
        # Log the chart creation
        log_audit(db, current_user_id, "export", "chart", f"Created {data_type} bar chart")
        
        return {
            "chart_type": "bar",
            "data_type": data_type,
            "image_base64": chart_base64,
            "created_at": datetime.utcnow().isoformat()
        }
    else:
        raise HTTPException(status_code=404, detail="No data available for chart")

@app.get("/charts/pie")
async def create_pie_chart(
    data_type: str = Query(..., description="Type of data to chart"),
    class_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user_id: int = Depends(RoleChecker(["admin", "hod", "teacher"]))
):
    """Create a pie chart"""
    current_user = db.query(User).filter(User.id == current_user_id).first()
    
    # Prepare data based on type
    if data_type == "role_distribution":
        query = db.query(User)
        if current_user.role == "hod":
            query = query.filter(User.department_id == current_user.department_id)
        
        users = query.all()
        data = pd.DataFrame([{
            'Role': user.role,
            'Count': 1
        } for user in users])
        
        if not data.empty:
            data = data.groupby('Role').size().reset_index(name='Count')
            chart_base64 = create_chart(data, 'pie', 'User Role Distribution')
        else:
            chart_base64 = None
    
    elif data_type == "department_distribution":
        query = db.query(Department)
        departments = query.all()
        data = pd.DataFrame([{
            'Department': dept.name,
            'Count': db.query(User).filter(User.department_id == dept.id).count()
        } for dept in departments])
        
        if not data.empty:
            chart_base64 = create_chart(data, 'pie', 'Department Distribution')
        else:
            chart_base64 = None
    
    else:
        raise HTTPException(status_code=400, detail="Invalid data type")
    
    if chart_base64:
        # Log the chart creation
        log_audit(db, current_user_id, "export", "chart", f"Created {data_type} pie chart")
        
        return {
            "chart_type": "pie",
            "data_type": data_type,
            "image_base64": chart_base64,
            "created_at": datetime.utcnow().isoformat()
        }
    else:
        raise HTTPException(status_code=404, detail="No data available for chart")

@app.get("/health")
async def health_check():
    return {"status": "healthy", "service": "exports"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8020)